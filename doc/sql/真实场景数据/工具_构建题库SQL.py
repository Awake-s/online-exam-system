# -*- coding: utf-8 -*-
"""题库构建器：纯 Excel 驱动 → 生成 02_题库.sql
数据源：doc/题库Excel/课程300题/*.xlsx（14 学科，唯一真相源）
执行：python 工具_构建题库SQL.py
"""
import json, os, re
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = r"d:\Java Projects\在线考试系统\doc\题库Excel\课程300题"
OUT_SQL = os.path.join(ROOT, "02_题库.sql")

QTYPE_MAP = {"单选": 1, "多选": 2, "判断": 3, "填空": 4, "简答": 5}
DIFF_MAP = {"简单": 1, "中等": 2, "困难": 3}

# 14 个 Excel 文件 → (学科ID, 学科名, 创建者教师ID)
EXCEL_TO_SUBJECT = {
    # 学科基础 / 专业核心（8 学科 × ~90 题）
    "高等数学.xlsx":       (100, "高等数学（一）", 100),
    "数据结构.xlsx":       (104, "数据结构", 110),
    "计算机组成原理.xlsx": (106, "计算机组成原理", 112),
    "计算机网络.xlsx":     (107, "计算机网络原理及工程应用", 114),
    "数据库原理.xlsx":     (108, "数据库原理", 116),
    "Java程序设计.xlsx":   (109, "Java EE 开发技术基础", 117),
    "软件工程导论.xlsx":   (110, "软件工程", 118),
    "人工智能导论.xlsx":   (112, "人工智能基础", 114),
    # 通识 / 补充学科（6 学科 × 60 题，v2.3 从 JSON 转为 Excel）
    "大学英语.xlsx":       (101, "大学英语", 104),
    "线性代数.xlsx":       (102, "线性代数", 107),
    "离散结构.xlsx":       (103, "离散结构", 109),
    "操作系统.xlsx":       (105, "操作系统", 112),
    "编译原理.xlsx":       (111, "编译原理", 109),
    "日语精读.xlsx":       (113, "日语精读", 106),
}


def sql_esc(s):
    if s is None: return "NULL"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "''") + "'"


def normalize(row, src):
    if not row or len(row) < 10 or row[0] is None or row[1] is None: return None
    qt = QTYPE_MAP.get(str(row[0]).strip())
    if qt is None: return None
    content = str(row[1]).strip()
    if not content: return None
    a, b, c, d = row[2:6]
    ans = "" if row[6] is None else str(row[6]).strip()
    ana = ("" if row[7] is None else str(row[7]).strip()) or src
    try: score = float(row[8]) if row[8] is not None else 2.0
    except: score = 2.0
    diff = DIFF_MAP.get(str(row[9]).strip() if row[9] else "", 1)

    opts = None
    if qt in (1, 2):
        d_ = {k: str(v).strip() for k, v in zip("ABCD", (a, b, c, d)) if v is not None and str(v).strip()}
        if len(d_) >= 2:
            opts = json.dumps(d_, ensure_ascii=False)
    if qt == 1:
        ans = ans.upper()
        if ans not in "ABCD": return None
    elif qt == 2:
        cleaned = re.sub(r"[\s,，、]", "", ans).upper()
        chars = sorted(set(c for c in cleaned if c in "ABCD"))
        if len(chars) < 2: return None
        ans = "".join(chars)
    elif qt == 3:
        if ans in ("对", "正确", "T", "True", "true", "√", "Y", "y"): ans = "正确"
        elif ans in ("错", "错误", "F", "False", "false", "×", "N", "n"): ans = "错误"
        else: return None
    else:
        if not ans: return None
    return {"qtype": qt, "content": content, "options": opts, "answer": ans,
            "analysis": ana, "score": score, "difficulty": diff}


def load_excel():
    bank = {}
    for fn, (sid, sn, cid) in EXCEL_TO_SUBJECT.items():
        full = os.path.join(EXCEL_DIR, fn)
        if not os.path.exists(full):
            print(f"[WARN] 未找到 {fn}，跳过")
            continue
        wb = load_workbook(full, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        qs = []
        for r in rows[1:]:
            q = normalize(r, f"来源：{fn}")
            if q:
                q["creator_id"] = cid
                qs.append(q)
        bank[sid] = qs
        print(f"[OK] {fn} -> subject {sid} ({sn}): {len(qs)} 题")
    return bank


def render(bank):
    L = []
    L.append("-- " + "=" * 60)
    L.append("-- 在线考试系统 · 题库 v2.5 纯 Excel 真实题库版")
    L.append("-- 数据源：doc/题库Excel/课程300题/*.xlsx（14 学科唯一真相源）")
    L.append("-- 构建脚本：python 工具_构建题库SQL.py")
    L.append("-- " + "=" * 60)
    L.append("USE online_exam_system;")
    L.append("SET NAMES utf8mb4;")
    L.append("SET FOREIGN_KEY_CHECKS = 0;")
    L.append("START TRANSACTION;")
    L.append("DELETE FROM exam_question WHERE id >= 100;")
    L.append("")

    qid = 1000
    ranges = []
    name_lookup = {v[0]: v[1] for v in EXCEL_TO_SUBJECT.values()}

    for sid in sorted(bank.keys()):
        qs = bank[sid]
        if not qs: continue
        sn = name_lookup.get(sid, "?")
        start = qid
        L.append(f"-- subject {sid} {sn} ({len(qs)} 题, ID {start}-{start+len(qs)-1})")
        L.append("INSERT INTO exam_question (id, subject_id, question_type, content, options, answer, analysis, score, difficulty, creator_id) VALUES")
        rows = []
        for q in qs:
            rows.append(
                f"({qid}, {sid}, {q['qtype']}, {sql_esc(q['content'])}, "
                f"{sql_esc(q['options'])}, {sql_esc(q['answer'])}, "
                f"{sql_esc(q['analysis'])}, {q['score']:.2f}, {q['difficulty']}, {q['creator_id']})"
            )
            qid += 1
        L.append(",\n".join(rows) + ";")
        L.append("")
        ranges.append((sid, sn, start, qid - 1, len(qs)))

    L.append("SET FOREIGN_KEY_CHECKS = 1;")
    L.append("COMMIT;")
    L.append("")
    L.append("-- 验证断言")
    L.append(f"SELECT '题库总数' AS 指标, COUNT(*) AS 实际, {qid-1000} AS 预期 FROM exam_question WHERE id >= 1000")
    for sid, sn, s, e, cnt in ranges:
        L.append(f"UNION ALL SELECT '{sid} {sn}', COUNT(*), {cnt} FROM exam_question WHERE subject_id={sid} AND id BETWEEN {s} AND {e}")
    L.append(";")
    L.append("")
    L.append("SELECT s.id AS 学科ID, s.subject_name AS 学科,")
    L.append("  SUM(CASE WHEN q.question_type=1 THEN 1 ELSE 0 END) AS 单选,")
    L.append("  SUM(CASE WHEN q.question_type=2 THEN 1 ELSE 0 END) AS 多选,")
    L.append("  SUM(CASE WHEN q.question_type=3 THEN 1 ELSE 0 END) AS 判断,")
    L.append("  SUM(CASE WHEN q.question_type=4 THEN 1 ELSE 0 END) AS 填空,")
    L.append("  SUM(CASE WHEN q.question_type=5 THEN 1 ELSE 0 END) AS 简答,")
    L.append("  COUNT(*) AS 总数 FROM edu_subject s LEFT JOIN exam_question q ON q.subject_id=s.id AND q.id>=1000")
    L.append("WHERE s.id >= 100 GROUP BY s.id, s.subject_name ORDER BY s.id;")
    return "\n".join(L) + "\n"


if __name__ == "__main__":
    bank = load_excel()
    sql = render(bank)
    with open(OUT_SQL, "w", encoding="utf-8") as f:
        f.write(sql)
    total = sum(len(v) for v in bank.values())
    print(f"\n[OUT] 写入 {OUT_SQL}")
    print(f"[STAT] 总题数: {total}（覆盖 {len(bank)} 学科）")
