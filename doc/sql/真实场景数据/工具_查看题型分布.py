# -*- coding: utf-8 -*-
"""
调研 doc/题库Excel 下 14 个学科 Excel 文件的题型分布
输出每个学科每种题型的题目数量,以及难度分布
"""
import os
import sys
from collections import Counter, defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少 openpyxl,请运行: pip install openpyxl")
    sys.exit(1)

EXCEL_DIR = r"d:\Java Projects\在线考试系统\doc\题库Excel\课程300题"

# 题型字段在中文 / 数字 之间的可能值
def normalize_qtype(v):
    if v is None:
        return "未知"
    s = str(v).strip()
    if s in ("1", "单选", "单选题"):
        return "单选题"
    if s in ("2", "多选", "多选题"):
        return "多选题"
    if s in ("3", "判断", "判断题"):
        return "判断题"
    if s in ("4", "填空", "填空题"):
        return "填空题"
    if s in ("5", "简答", "简答题", "问答", "问答题"):
        return "简答题"
    return f"其他({s})"

def normalize_difficulty(v):
    if v is None:
        return "未知"
    s = str(v).strip()
    if s in ("1", "简单", "易"):
        return "简单"
    if s in ("2", "中等", "中"):
        return "中等"
    if s in ("3", "困难", "难"):
        return "困难"
    return f"其他({s})"

def inspect_one(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    header = [str(c).strip() if c else "" for c in rows[0]]
    # 找到 题型 / 难度 字段索引
    qtype_idx = next((i for i, h in enumerate(header) if "题型" in h or "类型" in h), None)
    diff_idx = next((i for i, h in enumerate(header) if "难度" in h), None)
    if qtype_idx is None:
        return {"error": f"未找到题型列, 表头={header}"}
    
    qtype_counter = Counter()
    qtype_diff = defaultdict(Counter)
    total = 0
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        qt = normalize_qtype(row[qtype_idx])
        df = normalize_difficulty(row[diff_idx]) if diff_idx is not None else "未知"
        qtype_counter[qt] += 1
        qtype_diff[qt][df] += 1
        total += 1
    return {
        "header": header,
        "total": total,
        "qtype": dict(qtype_counter),
        "qtype_diff": {k: dict(v) for k, v in qtype_diff.items()},
    }

def main():
    files = sorted([f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx") and not f.startswith("~")])
    print(f"\n[发现 {len(files)} 个 Excel 文件]\n")
    
    grand_total = 0
    qtype_grand = Counter()
    
    print(f"{'学科':<20} {'总题数':>6} | {'单选':>5} {'多选':>5} {'判断':>5} {'填空':>5} {'简答':>5} | {'简单':>5} {'中等':>5} {'困难':>5}")
    print("-" * 110)
    
    for fname in files:
        path = os.path.join(EXCEL_DIR, fname)
        result = inspect_one(path)
        if result is None or "error" in result:
            print(f"{fname:<20} ERROR: {result.get('error', '未知')}")
            continue
        subject = fname.replace(".xlsx", "")
        qt = result["qtype"]
        # 难度统计 across all types
        diff_total = Counter()
        for _, dmap in result["qtype_diff"].items():
            for d, c in dmap.items():
                diff_total[d] += c
        
        print(f"{subject:<20} {result['total']:>6} | "
              f"{qt.get('单选题', 0):>5} {qt.get('多选题', 0):>5} {qt.get('判断题', 0):>5} "
              f"{qt.get('填空题', 0):>5} {qt.get('简答题', 0):>5} | "
              f"{diff_total.get('简单', 0):>5} {diff_total.get('中等', 0):>5} {diff_total.get('困难', 0):>5}")
        
        grand_total += result["total"]
        for k, v in qt.items():
            qtype_grand[k] += v
    
    print("-" * 110)
    print(f"{'合计':<20} {grand_total:>6} | "
          f"{qtype_grand.get('单选题', 0):>5} {qtype_grand.get('多选题', 0):>5} "
          f"{qtype_grand.get('判断题', 0):>5} {qtype_grand.get('填空题', 0):>5} "
          f"{qtype_grand.get('简答题', 0):>5}")

if __name__ == "__main__":
    main()
