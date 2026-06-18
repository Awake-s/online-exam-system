# -*- coding: utf-8 -*-
"""
题库扩容工具(v2.5)

工作流:
  1. 读基线 Excel(doc/题库Excel/课程300题/<subject>.xlsx)
  2. 加载 题库增量数据/<module>.py 的 INCREMENTAL 增量数据
  3. 按题目内容做去重合并
  4. 就地写回同一目录的 <subject>.xlsx
  5. 报告每题型最终数量 + 是否达标(目标 单选100 多选60 判断60 填空50 简答30 = 300)

用法:
  python 工具_扩容题库.py                  # 处理已注册的所有学科
  python 工具_扩容题库.py --subject 数据结构  # 只处理一个学科
  python 工具_扩容题库.py --check           # 只检查不写文件
"""
import argparse
import importlib
import os
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("[ERROR] 缺少 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)

# 目录配置
SCRIPT_DIR = Path(__file__).resolve().parent
# 基线 Excel 与输出目录均指向 课程300题/(运行后会就地刷新)
EXCEL_BASE_DIR = Path(r"d:\Java Projects\在线考试系统\doc\题库Excel\课程300题")
EXCEL_OUT_DIR = EXCEL_BASE_DIR

# 学科注册表：(Excel 文件名(不含扩展), 增量模块名)
# 每完成一个学科,取消注释对应行即可
SUBJECTS = [
    ("数据结构", "数据结构"),
    ("高等数学", "高等数学"),
    ("操作系统", "操作系统"),
    ("计算机网络", "计算机网络"),
    ("数据库原理", "数据库原理"),
    ("Java程序设计", "Java程序设计"),
    ("计算机组成原理", "计算机组成原理"),
    ("软件工程导论", "软件工程导论"),
    ("人工智能导论", "人工智能"),
    ("大学英语", "大学英语"),
    ("线性代数", "线性代数"),
    ("离散结构", "离散结构"),
    ("编译原理", "编译原理"),
    ("日语精读", "日语精读"),
]

# 目标分布(每学科)
TARGET_DISTRIBUTION = {"单选": 100, "多选": 60, "判断": 60, "填空": 50, "简答": 30}
TARGET_TOTAL = sum(TARGET_DISTRIBUTION.values())  # 300

HEADERS = ["题型", "题目内容", "选项A", "选项B", "选项C", "选项D", "答案", "解析", "分值", "难度"]


def load_baseline(subject_name: str):
    """读基线 Excel,返回行列表(不含表头),每行是 tuple"""
    path = EXCEL_BASE_DIR / f"{subject_name}.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"基线 Excel 不存在: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳过表头
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # 规范为 10 列(填充 None)
        row = tuple(list(row) + [None] * (10 - len(row)))[:10]
        rows.append(row)
    return rows


def load_incremental(module_name: str):
    """从 题库增量数据/<module_name>.py 加载 INCREMENTAL"""
    full_name = f"题库增量数据.{module_name}"
    sys.path.insert(0, str(SCRIPT_DIR))
    try:
        mod = importlib.import_module(full_name)
        importlib.reload(mod)
        return list(getattr(mod, "INCREMENTAL", []))
    except ModuleNotFoundError:
        print(f"[WARN] 增量模块未找到: {full_name},跳过")
        return []


def normalize_row(row):
    """规范化一行:strip 题目内容、统一答案格式(去空格)"""
    if not row:
        return row
    row = list(row)
    # 题目内容 strip
    if row[1] is not None:
        row[1] = str(row[1]).strip()
    # 答案去空格(尤其多选题的 "A, B, C" → "A,B,C")
    if row[6] is not None:
        s = str(row[6]).strip()
        if "," in s and any(c.isupper() for c in s):
            s = ",".join(p.strip() for p in s.split(","))
        row[6] = s
    return tuple(row)


def merge_dedupe(baseline_rows, incremental_rows):
    """按题目内容(第 2 列)去重,基线优先保留"""
    seen = set()
    out = []
    for row in baseline_rows + incremental_rows:
        row = normalize_row(row)
        key = (row[0], (row[1] or "").strip())  # (题型, 题目内容)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def count_by_qtype(rows):
    """按题型计数"""
    result = {k: 0 for k in TARGET_DISTRIBUTION}
    for r in rows:
        qt = str(r[0]).strip() if r[0] else ""
        if qt in result:
            result[qt] += 1
    return result


def write_excel(subject_name: str, rows):
    """写到 v2.5_expanded/ 目录"""
    EXCEL_OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = EXCEL_OUT_DIR / f"{subject_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        ws.append(list(r))
    wb.save(out_path)
    return out_path


def report(subject_name: str, baseline_count, incremental_count, final_rows, dry_run=False):
    """输出报告"""
    final_by_qt = count_by_qtype(final_rows)
    total = sum(final_by_qt.values())
    print(f"\n===== {subject_name} =====")
    print(f"基线:{baseline_count} | 增量:{incremental_count} | 合并去重后:{len(final_rows)}")
    print(f"目标 = 单选100 多选60 判断60 填空50 简答30 (合计 300)")
    print(f"  {'题型':<6} {'当前':>6} {'目标':>6} {'缺口':>6} {'状态':>4}")
    for qt, target in TARGET_DISTRIBUTION.items():
        cur = final_by_qt.get(qt, 0)
        gap = target - cur
        status = "OK" if cur >= target else "缺"
        print(f"  {qt:<6} {cur:>6} {target:>6} {gap:>6} {status:>4}")
    overall = "达标" if total >= TARGET_TOTAL else f"差 {TARGET_TOTAL - total} 题"
    print(f"  合计: {total} / {TARGET_TOTAL}  → {overall}")
    if dry_run:
        print(f"  [dry-run] 未写文件")


def process_subject(subject_name: str, module_name: str, dry_run: bool = False):
    try:
        baseline = load_baseline(subject_name)
    except FileNotFoundError as e:
        print(f"[ERROR] {e}")
        return False
    incremental = load_incremental(module_name)
    final_rows = merge_dedupe(baseline, incremental)
    report(subject_name, len(baseline), len(incremental), final_rows, dry_run=dry_run)
    if not dry_run:
        out_path = write_excel(subject_name, final_rows)
        print(f"  → 已写: {out_path}")
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--subject", help="只处理指定学科 (Excel 文件名,不含扩展)")
    parser.add_argument("--check", action="store_true", help="只检查不写文件")
    args = parser.parse_args()

    subjects_to_run = SUBJECTS
    if args.subject:
        subjects_to_run = [s for s in SUBJECTS if s[0] == args.subject]
        if not subjects_to_run:
            print(f"[ERROR] 学科 '{args.subject}' 未在 SUBJECTS 注册表中")
            sys.exit(1)

    print(f"v2.5 题库扩容 - 将处理 {len(subjects_to_run)} 个学科\n")
    for subject_name, module_name in subjects_to_run:
        process_subject(subject_name, module_name, dry_run=args.check)


if __name__ == "__main__":
    main()
